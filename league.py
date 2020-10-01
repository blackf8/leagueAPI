#league.py
#Approaches
    #existing code?/method that pro teams use
    #Keep track of everything

#track edward for vision (pink wards)
#mike days gone before he plays master yi (cstar specific)
import requests
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime


#jan 1 1970 midnight, this method converts epochtime to real time
def epochConvert(epochTime):
    epochTime = epochTime / 1000
    return datetime.fromtimestamp(epochTime)

#This method takes a champion id and returns the chamion name.
def champJsonIdentifier(champId):
    url = "http://ddragon.leagueoflegends.com/cdn/10.18.1/data/en_US/champion.json"
    response = requests.get(url)
    response.encoding = 'utf8'
    champJson = response.json()
    #champJson['data'].keys() list of champions
    #champJson['data']['Aatrox']['key']    specific champion key.
    for champ in champJson['data'].keys():
        if(champId == champJson['data'][champ]['key']):
            return champ

    return None

def matchHistory(workbook, playerData,accountId, SetDevKey):
    matchHisUrl = "https://na1.api.riotgames.com/lol/match/v4/matchlists/by-account/" + accountId + "?api_key=" + SetDevKey
    #https://na1.api.riotgames.com/lol/match/v4/matchlists/by-account/6uecamdtVoCgHZd8EzJmiXlDUkT6j_RyABUBLNrHj8cdeYI?api_key=RGAPI-a1c19a20-476d-4b68-ab76-3df0c31c90b1
    response = requests.get(matchHisUrl)
    response.encoding = 'utf8'
    json = response.json()

    print(json.keys())
    print(json['startIndex'],json['endIndex'],json['totalGames'])
    print(json['matches'][0])
    return None

def openpyxl(playerData):
    cwd = os.getcwd()
    listDir = os.listdir(cwd)
    if("league.xlsx" in listDir):
        workbook = load_workbook(filename="league.xlsx")
    else:
        workbook = Workbook()
    sheet = workbook.active
    counter = 65
    for element in playerData.keys():
        sheet[chr(counter) + "2" ] = element
        sheet[chr(counter) + "3"] = playerData[element]
        counter  = counter + 1
    workbook.save(filename="league.xlsx")
    return workbook
# Method: PlayerJson
# Details: Uses an url to make an api call and retrieves json data
# @param url: A string containing a url to a playerData json
# @return data: A json object holding player information
def PlayerJson(url):
    response = requests.get(url) #'https://google.com/'
    print(response)
    response.encoding = 'utf-8'
    data = response.json() #.content returns the byte encoding but we need it in utf8 formatting
    #print(data.keys())
    return data

# Method: PlayerIntro
# Details: Ask the user for basic information including a riot dev key/ign.
# @return url: A string containing a url to a playerData json
def PlayerIntro(key):

    #key = '' #input()
    print(f"Dev Key: {key}")

    SetIGN = 'Galactic Emperor' #input()
    print(f"Player Name: {SetIGN}")

    url = "https://na1.api.riotgames.com/lol/summoner/v4/summoners/by-name/" + SetIGN + "?api_key=" + key;
    return url

# MethodName: mains1
# Details: The main method
def main():
    SetDevKey = 'RGAPI-5fcdbea0-dedc-4e87-8aa3-04110c5c6fc9'
    url = PlayerIntro(SetDevKey)
    playerData = PlayerJson(url)
    for playerelement in playerData.keys():
        print(f"{playerelement}: {playerData[playerelement]}")

    workbook = openpyxl(playerData)
    matchHistory(workbook, playerData,playerData['accountId'], SetDevKey)
    champName = champJsonIdentifier("76")
    print(champName)
    print(epochConvert(1599948239606))

main()
